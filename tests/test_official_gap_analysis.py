from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook

from agent_review.eval.official_gap_analysis import (
    analyze_official_vs_report,
    load_official_review_baseline,
    parse_reviewer_report_titles,
    render_official_gap_markdown,
)
from agent_review.official_rule_registry import OFFICIAL_RULE_BY_NAME, OFFICIAL_RULE_PROFILES, build_title_synonym_groups


def test_load_official_review_baseline_reads_xlsx_rows(tmp_path: Path) -> None:
    workbook_path = tmp_path / "official.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(["埋点原文及页码", "审查点", "审查场景", "审查规则", "审查类别"])
    sheet.append(
        [
            "埋点原文：个体工商户等其他组织形式不得参与投标。埋点页码：19",
            "不得非法限定供应商所有制形式、组织形式、注册地或所在地",
            "不得限定供应商组织形式",
            "不得限定供应商组织形式",
            "一、资格公平性",
        ]
    )
    workbook.save(workbook_path)

    baseline = load_official_review_baseline(workbook_path)

    assert baseline.sheet_name == "Sheet1"
    assert len(baseline.items) == 1
    assert baseline.items[0].anchor_text == "个体工商户等其他组织形式不得参与投标。"
    assert baseline.items[0].page_hint == "19"


def test_parse_reviewer_report_titles_extracts_formal_titles(tmp_path: Path) -> None:
    report_path = tmp_path / "reviewer_report.md"
    report_path.write_text(
        "\n".join(
            [
                "**招标文件合规审查意见书**",
                "**1. 资格条件可能缺乏履约必要性或带有歧视性门槛**",
                "**2. 行业错配评分项被纳入评审**",
            ]
        ),
        encoding="utf-8",
    )

    titles = parse_reviewer_report_titles(report_path)

    assert titles == [
        "资格条件可能缺乏履约必要性或带有歧视性门槛",
        "行业错配评分项被纳入评审",
    ]


def test_analyze_official_vs_report_identifies_partial_and_missed_items(tmp_path: Path) -> None:
    workbook_path = tmp_path / "official.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["埋点原文及页码", "审查点", "审查场景", "审查规则", "审查类别"])
    sheet.append(
        [
            "埋点原文：投标人须具备全国科技型中小企业证书。埋点页码：19",
            "不得将供应商规模条件、股权结构、年限设置为资格条件",
            "资格公平性",
            "不得将营业收入的隐性限制证书设置为资格条件",
            "一、资格公平性",
        ]
    )
    sheet.append(
        [
            "埋点原文：4、投标人的注册资本200万以上加20分。埋点页码：15",
            "依法设定评审因素",
            "评审规则合规性",
            "不得将供应商的注册资本设定为评审因素",
            "三、评审规则合规性",
        ]
    )
    workbook.save(workbook_path)

    report_path = tmp_path / "reviewer_report.md"
    report_path.write_text(
        "\n".join(
            [
                "**招标文件合规审查意见书**",
                "**1. 资格条件可能缺乏履约必要性或带有歧视性门槛**",
                "**2. 专利要求**",
            ]
        ),
        encoding="utf-8",
    )
    trace_path = tmp_path / "trace.json"
    trace_path.write_text(
        json.dumps(
            {
                "requested_mode": "enhanced",
                "final_mode": "enhanced",
                "llm_enhanced": True,
                "warnings": ["llm_scenario_review 未生效：mocked"],
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    analysis = analyze_official_vs_report(workbook_path, report_path, trace_path)

    assert analysis.official_item_count == 2
    assert analysis.partial_match_count == 1
    assert analysis.missed_count == 1
    assert analysis.partial_match_items[0].matched_report_title == "资格条件可能缺乏履约必要性或带有歧视性门槛"
    assert analysis.missed_items[0].root_cause == "缺少官方规则族与细粒度事实槽位"
    assert "专利要求" in analysis.false_positive_titles


def test_render_official_gap_markdown_contains_sections(tmp_path: Path) -> None:
    workbook_path = tmp_path / "official.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["埋点原文及页码", "审查点", "审查场景", "审查规则", "审查类别"])
    sheet.append(
        [
            "埋点原文：本项目建设周期 38 个月。埋点页码：29",
            "合理设置合同履行期限",
            "内容规范性",
            "服务合同履行期限不得超过36个月",
            "四、内容规范性",
        ]
    )
    workbook.save(workbook_path)

    report_path = tmp_path / "reviewer_report.md"
    report_path.write_text("**招标文件合规审查意见书**", encoding="utf-8")

    analysis = analyze_official_vs_report(workbook_path, report_path)
    markdown = render_official_gap_markdown(analysis)

    assert "# 官方结果对比分析" in markdown
    assert "## 漏检" in markdown
    assert "合理设置合同履行期限" in markdown


def test_official_rule_registry_contains_19_rules() -> None:
    assert len(OFFICIAL_RULE_PROFILES) == 19
    assert OFFICIAL_RULE_BY_NAME["不得将供应商的资产总额设定为评审因素"].authority_binding_ids == ["AUTH-RP-SCORE-014-001"]


def test_official_title_synonym_groups_cover_price_method_rule() -> None:
    groups = build_title_synonym_groups()
    assert "综合评分法价格分未采用低价优先法" in groups["采用综合评分法评标的，价格分（必须）采用低价优先法"]
    assert "依法设定价格分值" in groups["采用综合评分法评标的，价格分（必须）采用低价优先法"]


def test_partial_match_does_not_fall_back_to_generic_scoring_title(tmp_path: Path) -> None:
    workbook_path = tmp_path / "official.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["埋点原文及页码", "审查点", "审查场景", "审查规则", "审查类别"])
    sheet.append(
        [
            "埋点原文：投标人的资产总额达到5000万元以上的，得5分。埋点页码：12",
            "依法设定评审因素",
            "评审规则合规性",
            "不得将供应商的资产总额设定为评审因素",
            "三、评审规则合规性",
        ]
    )
    workbook.save(workbook_path)

    report_path = tmp_path / "reviewer_report.md"
    report_path.write_text(
        "\n".join(
            [
                "**招标文件合规审查意见书**",
                "**1. 净利润或利润被设为评分因素**",
            ]
        ),
        encoding="utf-8",
    )

    analysis = analyze_official_vs_report(workbook_path, report_path)

    assert analysis.partial_match_count == 0
    assert analysis.missed_count == 1


def test_official_gap_analysis_maps_requirement_and_technical_bundle_into_partial_matches(tmp_path: Path) -> None:
    workbook_path = tmp_path / "official_bundle.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["埋点原文及页码", "审查点", "审查场景", "审查规则", "审查类别"])
    rows = [
        ("埋点原文：深圳市医疗器械行业同类项目业绩不少于2个。埋点页码：11", "不得将特定行业、特定行政区划、特定金额的业绩、奖项作为资格条件", "资格公平性", "不得将特定行业、特定行政区划、特定金额的业绩作为资格条件", "一、资格公平性"),
        ("埋点原文：验收时产生的第三方检测费用由中标人承担，无论检测结果是否合格。埋点页码：18", "不得要求中标人承担检测费用", "需求合规性", "不得将检测费用与检测结果合格性挂钩", "二、需求合规性"),
        ("埋点原文：产品须符合GB/T 99999-2024。埋点页码：22", "检测报告/检测标准合规性", "需求合规性", "不得使用不存在的检测标准", "二、需求合规性"),
        ("埋点原文：产品响应时间：100ms。埋点页码：23", "技术参数的区间设置合规性", "需求合规性", "不得缺失“技术参数区间说明”", "二、需求合规性"),
        ("埋点原文：设备重量≤500kg(不允许正偏离)；设备重量允许±10%偏差。埋点页码：24", "技术参数的区间设置合规性", "需求合规性", "不得存在区间说明冲突", "二、需求合规性"),
        ("埋点原文：系统应具有良好的操作体验，界面友好美观。埋点页码：25", "依法设定评审因素", "评审规则合规性", "不得使用主观描述来要求供应商所投产品", "三、评审规则合规性"),
    ]
    for row in rows:
        sheet.append(list(row))
    workbook.save(workbook_path)

    report_path = tmp_path / "reviewer_report.md"
    report_path.write_text(
        "\n".join(
            [
                "**招标文件合规审查意见书**",
                "**1. 资格业绩要求可能存在地域限定、行业口径过窄或与评分重复**",
                "**2. 第三方检测费用无论结果均由中标人承担**",
                "**3. 疑似使用不存在的技术标准**",
                "**4. 技术参数区间说明不足**",
                "**5. 同一技术参数区间说明冲突**",
                "**6. 技术要求存在主观描述**",
            ]
        ),
        encoding="utf-8",
    )

    analysis = analyze_official_vs_report(workbook_path, report_path)

    assert analysis.partial_match_count == 6
    assert analysis.missed_count == 0
