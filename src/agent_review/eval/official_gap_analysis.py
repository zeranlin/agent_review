from __future__ import annotations

import json
import re
import zipfile
from collections import Counter
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook

from ..models import OfficialGapAnalysis, OfficialGapItem, OfficialReviewBaseline, OfficialReviewItem
from ..official_rule_registry import OFFICIAL_RULE_BY_NAME, build_title_synonym_groups


REPORT_TITLE_PATTERN = re.compile(r"^\*\*\d+\.\s+(.+?)\*\*$", re.MULTILINE)
PAGE_HINT_PATTERN = re.compile(r"页码[:：]\s*([0-9Nn]+)")
ANCHOR_PATTERN = re.compile(r"埋点原文[:：]\s*(.+?)(?:埋点页码[:：]|$)", re.DOTALL)


TITLE_SYNONYM_GROUPS: dict[str, tuple[str, ...]] = build_title_synonym_groups()
GENERIC_REPORT_TITLES = {
    "资格条件可能缺乏履约必要性或带有歧视性门槛",
    "依法设定价格分值",
}
DETAIL_SHEET_CANDIDATES = (
    "埋点详情（优化结果验证）",
    "埋点详情",
)
HEADER_ALIASES = {
    "anchor_raw": ("埋点原文及页码",),
    "review_point": ("审查点",),
    "scenario": ("审查场景",),
    "rule_name": ("审查规则",),
    "category": ("审查类别",),
}
XML_NS = {
    "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "rel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "pkg": "http://schemas.openxmlformats.org/package/2006/relationships",
}


def load_official_review_baseline(path: str | Path) -> OfficialReviewBaseline:
    workbook_path = Path(path).expanduser().resolve()
    try:
        workbook = load_workbook(workbook_path, data_only=True)
        sheet = _select_openpyxl_sheet(workbook)
        rows = [[("" if cell is None else str(cell).strip()) for cell in row] for row in sheet.iter_rows(values_only=True)]
        sheet_name = sheet.title
    except Exception:
        sheet_name, rows = _load_xlsx_rows_without_styles(workbook_path)
    header_map = _resolve_header_map(rows[0] if rows else [])
    items: list[OfficialReviewItem] = []
    for row_index, row in enumerate(rows[1:], start=2):
        if not any(row):
            continue
        anchor_raw = _value_from_row(row, header_map, "anchor_raw")
        review_point = _value_from_row(row, header_map, "review_point")
        scenario = _value_from_row(row, header_map, "scenario")
        rule_name = _value_from_row(row, header_map, "rule_name")
        category = _value_from_row(row, header_map, "category")
        if not any([anchor_raw, review_point, scenario, rule_name, category]):
            continue
        anchor_text = _extract_anchor_text(anchor_raw)
        page_hint = _extract_page_hint(anchor_raw)
        items.append(
            OfficialReviewItem(
                row_index=row_index,
                anchor_text=anchor_text,
                page_hint=page_hint,
                review_point=review_point,
                scenario=scenario,
                rule_name=rule_name,
                category=category,
            )
        )
    return OfficialReviewBaseline(
        source_path=str(workbook_path),
        sheet_name=sheet_name,
        items=items,
    )


def _select_openpyxl_sheet(workbook) -> object:
    for candidate in DETAIL_SHEET_CANDIDATES:
        if candidate in workbook.sheetnames:
            return workbook[candidate]
    return workbook[workbook.sheetnames[0]]


def _resolve_header_map(header_row: list[str]) -> dict[str, int]:
    normalized = [str(cell).strip() for cell in header_row]
    mapping: dict[str, int] = {}
    for field, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            if alias in normalized:
                mapping[field] = normalized.index(alias)
                break
    if len(mapping) == len(HEADER_ALIASES):
        return mapping
    fallback_order = ["anchor_raw", "review_point", "scenario", "rule_name", "category"]
    return {field: index for index, field in enumerate(fallback_order)}


def _value_from_row(row: list[str], header_map: dict[str, int], field: str) -> str:
    index = header_map.get(field, -1)
    if index < 0 or index >= len(row):
        return ""
    return str(row[index]).strip()


def _load_xlsx_rows_without_styles(workbook_path: Path) -> tuple[str, list[list[str]]]:
    with zipfile.ZipFile(workbook_path) as archive:
        shared_strings = _read_shared_strings(archive)
        workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
        rel_root = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        sheet_targets = {
            rel.attrib.get("Id", ""): rel.attrib.get("Target", "")
            for rel in rel_root.findall("pkg:Relationship", XML_NS)
        }
        sheets: list[tuple[str, str]] = []
        for sheet in workbook_root.findall("main:sheets/main:sheet", XML_NS):
            name = sheet.attrib.get("name", "")
            rel_id = sheet.attrib.get(f"{{{XML_NS['rel']}}}id", "")
            target = sheet_targets.get(rel_id, "")
            if target:
                sheets.append((name, f"xl/{target}"))
        selected_name, selected_target = _select_xml_sheet(sheets)
        rows = _read_sheet_rows_from_xml(archive, selected_target, shared_strings)
        return selected_name, rows


def _select_xml_sheet(sheets: list[tuple[str, str]]) -> tuple[str, str]:
    for candidate in DETAIL_SHEET_CANDIDATES:
        for name, target in sheets:
            if name == candidate:
                return name, target
    return sheets[0] if sheets else ("", "")


def _read_shared_strings(archive: zipfile.ZipFile) -> list[str]:
    if "xl/sharedStrings.xml" not in archive.namelist():
        return []
    root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
    strings: list[str] = []
    for item in root.findall("main:si", XML_NS):
        strings.append("".join(text.text or "" for text in item.iterfind(".//main:t", XML_NS)))
    return strings


def _read_sheet_rows_from_xml(
    archive: zipfile.ZipFile,
    sheet_path: str,
    shared_strings: list[str],
) -> list[list[str]]:
    if not sheet_path or sheet_path not in archive.namelist():
        return []
    root = ET.fromstring(archive.read(sheet_path))
    sheet_data = root.find("main:sheetData", XML_NS)
    if sheet_data is None:
        return []
    rows: list[list[str]] = []
    for row in sheet_data.findall("main:row", XML_NS):
        values: list[str] = []
        current_col = 1
        for cell in row.findall("main:c", XML_NS):
            ref = cell.attrib.get("r", "")
            target_col = _column_index_from_ref(ref) if ref else current_col
            while current_col < target_col:
                values.append("")
                current_col += 1
            values.append(_read_cell_value(cell, shared_strings))
            current_col += 1
        rows.append(values)
    return rows


def _column_index_from_ref(ref: str) -> int:
    letters = "".join(char for char in ref if char.isalpha())
    index = 0
    for char in letters:
        index = index * 26 + (ord(char.upper()) - ord("A") + 1)
    return index or 1


def _read_cell_value(cell, shared_strings: list[str]) -> str:
    cell_type = cell.attrib.get("t", "")
    if cell_type == "inlineStr":
        return "".join(text.text or "" for text in cell.iterfind(".//main:t", XML_NS)).strip()
    value_node = cell.find("main:v", XML_NS)
    if value_node is None or value_node.text is None:
        return ""
    raw = value_node.text
    if cell_type == "s":
        try:
            return shared_strings[int(raw)].strip()
        except (ValueError, IndexError):
            return raw.strip()
    return raw.strip()


def parse_reviewer_report_titles(report_path: str | Path) -> list[str]:
    report_text = Path(report_path).expanduser().resolve().read_text(encoding="utf-8")
    return [match.group(1).strip() for match in REPORT_TITLE_PATTERN.finditer(report_text)]


def analyze_official_vs_report(
    official_xlsx_path: str | Path,
    reviewer_report_path: str | Path,
    enhancement_trace_path: str | Path | None = None,
) -> OfficialGapAnalysis:
    baseline = load_official_review_baseline(official_xlsx_path)
    report_titles = parse_reviewer_report_titles(reviewer_report_path)
    llm_status = _load_llm_status(enhancement_trace_path)

    matched_items: list[OfficialGapItem] = []
    partial_match_items: list[OfficialGapItem] = []
    missed_items: list[OfficialGapItem] = []
    matched_report_titles: set[str] = set()

    for item in baseline.items:
        matched_title = _find_direct_match(item, report_titles)
        if matched_title:
            matched_report_titles.add(matched_title)
            matched_items.append(
                OfficialGapItem(
                    row_index=item.row_index,
                    anchor_text=item.anchor_text,
                    page_hint=item.page_hint,
                    review_point=item.review_point,
                    rule_name=item.rule_name,
                    category=item.category,
                    matched_report_title=matched_title,
                    gap_stage="none",
                    gap_type="matched",
                    root_cause="主链已命中",
                    recommendation="保持当前规则与 formal 输出链。",
                )
            )
            continue

        partial_title = _find_partial_match(item, report_titles)
        if partial_title:
            matched_report_titles.add(partial_title)
            partial_match_items.append(
                OfficialGapItem(
                    row_index=item.row_index,
                    anchor_text=item.anchor_text,
                    page_hint=item.page_hint,
                    review_point=item.review_point,
                    rule_name=item.rule_name,
                    category=item.category,
                    matched_report_title=partial_title,
                    gap_stage="rule/report",
                    gap_type="partial_match",
                    root_cause=_infer_partial_root_cause(item),
                    recommendation=_recommendation_for_item(item, partial=True),
                )
            )
            continue

        missed_items.append(
            OfficialGapItem(
                row_index=item.row_index,
                anchor_text=item.anchor_text,
                page_hint=item.page_hint,
                review_point=item.review_point,
                rule_name=item.rule_name,
                category=item.category,
                gap_stage=_infer_missed_stage(item),
                gap_type="missed",
                root_cause=_infer_missed_root_cause(item, llm_status),
                recommendation=_recommendation_for_item(item, partial=False),
            )
        )

    false_positive_titles = [title for title in report_titles if title not in matched_report_titles]
    root_cause_counts = Counter(
        item.root_cause for item in [*partial_match_items, *missed_items] if item.root_cause
    )

    return OfficialGapAnalysis(
        source_xlsx_path=str(Path(official_xlsx_path).expanduser().resolve()),
        source_report_path=str(Path(reviewer_report_path).expanduser().resolve()),
        official_item_count=len(baseline.items),
        matched_count=len(matched_items),
        partial_match_count=len(partial_match_items),
        missed_count=len(missed_items),
        matched_items=matched_items,
        partial_match_items=partial_match_items,
        missed_items=missed_items,
        false_positive_titles=false_positive_titles,
        llm_status=llm_status,
        root_cause_counts=dict(root_cause_counts),
    )


def render_official_gap_markdown(analysis: OfficialGapAnalysis) -> str:
    lines = [
        "# 官方结果对比分析",
        "",
        f"- 官方点数: {analysis.official_item_count}",
        f"- 完全命中: {analysis.matched_count}",
        f"- 部分命中: {analysis.partial_match_count}",
        f"- 漏检: {analysis.missed_count}",
        "",
        "## 根因统计",
        "",
    ]
    if analysis.root_cause_counts:
        for key, value in sorted(analysis.root_cause_counts.items(), key=lambda item: (-item[1], item[0])):
            lines.append(f"- {key}: {value}")
    else:
        lines.append("- 当前无部分命中或漏检。")
    lines.append("")

    lines.extend(["## 部分命中", ""])
    if analysis.partial_match_items:
        for item in analysis.partial_match_items:
            lines.extend(_render_gap_item_lines(item))
    else:
        lines.append("- 无")
    lines.append("")

    lines.extend(["## 漏检", ""])
    if analysis.missed_items:
        for item in analysis.missed_items:
            lines.extend(_render_gap_item_lines(item))
    else:
        lines.append("- 无")
    lines.append("")

    lines.extend(["## 当前疑似误报/弱报", ""])
    if analysis.false_positive_titles:
        for title in analysis.false_positive_titles:
            lines.append(f"- {title}")
    else:
        lines.append("- 无")
    return "\n".join(lines)


def _render_gap_item_lines(item: OfficialGapItem) -> list[str]:
    return [
        f"### Row {item.row_index}",
        f"- 官方规则: {item.rule_name}",
        f"- 官方审查点: {item.review_point}",
        f"- 原文锚点: {item.anchor_text}",
        f"- 页码: {item.page_hint or '未提供'}",
        f"- 当前命中: {item.matched_report_title or '未命中'}",
        f"- 漏检阶段: {item.gap_stage}",
        f"- 根因: {item.root_cause}",
        f"- 建议: {item.recommendation}",
        "",
    ]


def _extract_anchor_text(raw: str) -> str:
    match = ANCHOR_PATTERN.search(raw or "")
    if match:
        return re.sub(r"\s+", " ", match.group(1)).strip(" ：:")
    return re.sub(r"\s+", " ", raw or "").strip()


def _extract_page_hint(raw: str) -> str:
    match = PAGE_HINT_PATTERN.search(raw or "")
    return match.group(1) if match else ""


def _find_direct_match(item: OfficialReviewItem, report_titles: list[str]) -> str:
    profile = OFFICIAL_RULE_BY_NAME.get(item.rule_name)
    if profile:
        preferred_titles = [title for title in profile.report_titles if title not in GENERIC_REPORT_TITLES]
        for expected_title in preferred_titles:
            for title in report_titles:
                if expected_title and (expected_title in title or title in expected_title):
                    return title
    rule_name = item.rule_name
    for title in report_titles:
        if rule_name and (rule_name in title or title in rule_name):
            return title
    anchor_text = item.anchor_text
    for title in report_titles:
        if anchor_text and any(token in title for token in _extract_signal_tokens(anchor_text)):
            return title
    return ""


def _find_partial_match(item: OfficialReviewItem, report_titles: list[str]) -> str:
    profile = OFFICIAL_RULE_BY_NAME.get(item.rule_name)
    if profile:
        generic_titles = [title for title in profile.report_titles if title in GENERIC_REPORT_TITLES]
        for expected_title in generic_titles:
            for title in report_titles:
                if expected_title and (expected_title in title or title in expected_title):
                    return title
        for title in report_titles:
            if any(candidate and (candidate in title or title in candidate) for candidate in profile.report_titles):
                return title
        for title in report_titles:
            hits = sum(1 for token in profile.matching_terms if token and token in title)
            if hits >= 2:
                return title
        return ""
    candidates = TITLE_SYNONYM_GROUPS.get(item.rule_name) or TITLE_SYNONYM_GROUPS.get(item.review_point) or ()
    for title in report_titles:
        if any(token in title for token in candidates):
            return title
    return ""


def _extract_signal_tokens(text: str) -> list[str]:
    return [token for token in re.split(r"[，。；、\s]+", text) if len(token) >= 4][:6]


def _infer_partial_root_cause(item: OfficialReviewItem) -> str:
    if any(token in item.rule_name for token in ["资格条件", "组织形式", "证书设置为资格条件"]):
        return "规则已部分覆盖但报告标题过泛"
    if "评审因素" in item.review_point or "评审因素" in item.rule_name:
        return "事实已抽取但缺少更细规则槽位"
    return "已被泛化规则吸收，未形成官方口径标题"


def _infer_missed_stage(item: OfficialReviewItem) -> str:
    if any(token in item.rule_name for token in ["价格分值", "履行期限", "工作日内完成资金支付"]):
        return "rule"
    if any(token in item.rule_name for token in ["注册资本", "营业收入", "利润", "股权结构", "经营年限"]):
        return "fact/rule"
    if any(token in item.rule_name for token in ["组织形式", "认证范围", "准入类"]):
        return "rule/formal"
    return "rule"


def _infer_missed_root_cause(item: OfficialReviewItem, llm_status: dict[str, object]) -> str:
    if any(token in item.rule_name for token in ["注册资本", "营业收入", "利润", "股权结构", "经营年限"]):
        return "缺少官方规则族与细粒度事实槽位"
    if any(token in item.rule_name for token in ["价格分值", "履行期限", "工作日内完成资金支付"]):
        return "缺少时限/比例类专门规则，当前主链未激活"
    if any(token in item.rule_name for token in ["组织形式", "认证范围", "准入类"]):
        return "文本已解析但未接入正式审查点或 formal 映射"
    if llm_status.get("scenario_review_failed"):
        return "需要场景化语义补偿，但本次 LLM scenario review 失败"
    return "规则库未覆盖，LLM 也未在关键节点补偿"


def _recommendation_for_item(item: OfficialReviewItem, *, partial: bool) -> str:
    if any(token in item.rule_name for token in ["注册资本", "营业收入", "利润", "股权结构", "经营年限"]):
        return "新增规模/财务/主体属性评分规则族，并补 metric_name/metric_category 事实槽位。"
    if any(token in item.rule_name for token in ["组织形式", "认证范围", "准入类"]):
        return "新增正式 ReviewPointContract 与报告标题映射，并补 formal 证据对齐。"
    if any(token in item.rule_name for token in ["价格分值", "履行期限", "工作日内完成资金支付"]):
        return "新增比例、期限、付款时限规则，并纳入地方定制法理绑定。"
    if partial:
        return "保留现有规则命中，同时把官方口径沉淀为更具体的报告簇。"
    return "补规则并让 LLM 仅在低置信度归类时介入。"


def _load_llm_status(path: str | Path | None) -> dict[str, object]:
    if path is None:
        return {}
    trace_path = Path(path).expanduser().resolve()
    if not trace_path.exists():
        return {}
    payload = json.loads(trace_path.read_text(encoding="utf-8"))
    warnings = payload.get("warnings", []) or []
    return {
        "requested_mode": payload.get("requested_mode", ""),
        "final_mode": payload.get("final_mode", ""),
        "llm_enhanced": payload.get("llm_enhanced", False),
        "scenario_review_failed": any("llm_scenario_review" in item for item in warnings),
        "warnings": warnings,
    }
